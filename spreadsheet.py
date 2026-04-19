from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import SETTINGS
from savant import MatchupSummary

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
TITLE_FILL = PatternFill("solid", fgColor="0F243E")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
GREEN_FILL = PatternFill("solid", fgColor="E2F0D9")
RED_FILL = PatternFill("solid", fgColor="FDE9E7")
ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")
TEAL_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN_GRAY = Side(style="thin", color="D9D9D9")
BORDER_BOTTOM = Border(bottom=Side(style="medium", color="7F7F7F"))


@dataclass(frozen=True)
class SheetRow:
    run_date: str
    game: str
    batter: str
    pitcher: str
    hits: int
    ab: int
    singles: int
    doubles: int
    triples: int
    home_runs: int
    last_ab_count: int
    last_results: str


def _summary_to_row(run_date: str, game_label: str, summary: MatchupSummary) -> SheetRow:
    return SheetRow(
        run_date=run_date,
        game=game_label,
        batter=summary.batter_name,
        pitcher=summary.pitcher_name,
        hits=summary.hits,
        ab=summary.ab,
        singles=summary.hit_breakdown.get("1B", 0),
        doubles=summary.hit_breakdown.get("2B", 0),
        triples=summary.hit_breakdown.get("3B", 0),
        home_runs=summary.hit_breakdown.get("HR", 0),
        last_ab_count=summary.last_ab_count,
        last_results=" | ".join(summary.last_results),
    )


class WorkbookExporter:
    def __init__(self, run_date: str, game_results: list[tuple[str, list[MatchupSummary]]]) -> None:
        self.run_date = run_date
        self.rows = [
            _summary_to_row(run_date, game_label, summary)
            for game_label, summaries in game_results
            for summary in summaries
        ]

    def export(self, output_path: str | Path) -> Path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        default = wb.active
        wb.remove(default)

        overview = wb.create_sheet("Overview")
        all_ws = wb.create_sheet("All Matchups")
        qualified_ws = wb.create_sheet("Qualified")

        self._build_overview(overview)
        self._build_data_sheet(all_ws, qualified_only=False)
        self._build_data_sheet(qualified_ws, qualified_only=True)

        wb.save(output_path)
        return output_path

    def _build_overview(self, ws) -> None:
        ws.sheet_view.showGridLines = False
        ws["A1"] = "Hits Matchup Bot"
        ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
        ws["A1"].fill = TITLE_FILL
        ws.merge_cells("A1:F1")

        ws["A3"] = "Run Settings"
        ws["A3"].font = Font(bold=True, color="FFFFFF")
        ws["A3"].fill = HEADER_FILL
        ws.merge_cells("A3:B3")

        generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
        settings_rows = [
            ("Run date", self.run_date),
            ("Generated (UTC)", generated_at),
            ("Min AB", SETTINGS.min_ab),
            ("Min Hits", SETTINGS.min_hits),
            ("Min Hit Rate", SETTINGS.min_hit_rate),
            ("Last AB Window", SETTINGS.last_ab_window),
            ("Search Mode", SETTINGS.search_mode),
            ("Max Hitters / Game", SETTINGS.max_hitters_per_game),
        ]
        for row_idx, (label, value) in enumerate(settings_rows, start=4):
            ws.cell(row=row_idx, column=1, value=label)
            cell = ws.cell(row=row_idx, column=2, value=value)
            cell.fill = INPUT_FILL
            cell.font = Font(color="0000FF")

        ws["D3"] = "Workbook KPIs"
        ws["D3"].font = Font(bold=True, color="FFFFFF")
        ws["D3"].fill = HEADER_FILL
        ws.merge_cells("D3:E3")

        kpis = [
            ("Total matchups", "=MAX(COUNTA('All Matchups'!B:B)-1,0)"),
            ("Qualified matchups", "=COUNTIF('All Matchups'!N:N,\"Yes\")"),
            ("Elite", '=COUNTIF(\'All Matchups\'!O:O,"Elite")'),
            ("Strong", '=COUNTIF(\'All Matchups\'!O:O,"Strong")'),
            ("Playable", '=COUNTIF(\'All Matchups\'!O:O,"Playable")'),
            ("Thin", '=COUNTIF(\'All Matchups\'!O:O,"Thin")'),
        ]
        for row_idx, (label, formula) in enumerate(kpis, start=4):
            ws.cell(row=row_idx, column=4, value=label)
            cell = ws.cell(row=row_idx, column=5)
            cell.value = formula
            cell.fill = TEAL_FILL
            cell.font = Font(bold=True)

        ws["A14"] = "Tier Rules"
        ws["A14"].font = Font(bold=True, color="FFFFFF")
        ws["A14"].fill = HEADER_FILL
        ws.merge_cells("A14:C14")
        tier_rules = [
            ("Elite", "5+ hits and .500+ hit rate"),
            ("Strong", "4+ hits and .400+ hit rate"),
            ("Playable", "Passes filter thresholds"),
            ("Thin", "Fails filter thresholds"),
        ]
        for row_idx, (tier, rule) in enumerate(tier_rules, start=15):
            ws.cell(row=row_idx, column=1, value=tier)
            ws.cell(row=row_idx, column=2, value=rule)

        ws["D14"] = "How to Read It"
        ws["D14"].font = Font(bold=True, color="FFFFFF")
        ws["D14"].fill = HEADER_FILL
        ws.merge_cells("D14:F14")
        notes = [
            "All Matchups = every batter vs probable starter matchup on the slate.",
            "Qualified = only rows that pass Min AB / Min Hits / Min Hit Rate.",
            "Hit Rate uses hits divided by AB only — not true OBP.",
            "Last Results shows recent AB results only, in most-recent-first order.",
        ]
        for row_idx, note in enumerate(notes, start=15):
            ws.cell(row=row_idx, column=4, value=note)

        for col, width in {"A": 22, "B": 18, "C": 4, "D": 22, "E": 16, "F": 52}.items():
            ws.column_dimensions[col].width = width

    def _build_data_sheet(self, ws, qualified_only: bool) -> None:
        ws.sheet_view.showGridLines = False
        headers = [
            "Run Date",
            "Game",
            "Batter",
            "Pitcher",
            "Hits",
            "AB",
            "Hit Rate",
            "1B",
            "2B",
            "3B",
            "HR",
            "Last AB Count",
            "Last Results",
            "Qualified",
            "Tier",
            "Matchup",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER_BOTTOM

        current_row = 2
        source_rows = self.rows
        if qualified_only:
            source_rows = [
                row for row in source_rows
                if row.ab >= SETTINGS.min_ab and row.hits >= SETTINGS.min_hits and (row.hits / row.ab if row.ab else 0.0) >= SETTINGS.min_hit_rate
            ]

        source_rows = sorted(
            source_rows,
            key=lambda row: (-(row.hits / row.ab if row.ab else 0.0), -row.hits, -row.ab, row.game, row.batter),
        )

        for row in source_rows:
            ws.cell(row=current_row, column=1, value=row.run_date)
            ws.cell(row=current_row, column=2, value=row.game)
            ws.cell(row=current_row, column=3, value=row.batter)
            ws.cell(row=current_row, column=4, value=row.pitcher)
            ws.cell(row=current_row, column=5, value=row.hits)
            ws.cell(row=current_row, column=6, value=row.ab)
            ws.cell(row=current_row, column=7, value=f"=IFERROR(E{current_row}/F{current_row},0)")
            ws.cell(row=current_row, column=8, value=row.singles)
            ws.cell(row=current_row, column=9, value=row.doubles)
            ws.cell(row=current_row, column=10, value=row.triples)
            ws.cell(row=current_row, column=11, value=row.home_runs)
            ws.cell(row=current_row, column=12, value=row.last_ab_count)
            ws.cell(row=current_row, column=13, value=row.last_results)
            ws.cell(row=current_row, column=14, value=(
                f'=IF(AND(F{current_row}>=Overview!$B$6,E{current_row}>=Overview!$B$7,G{current_row}>=Overview!$B$8),"Yes","No")'
            ))
            ws.cell(row=current_row, column=15, value=(
                f'=IF(AND(E{current_row}>=5,G{current_row}>=0.5),"Elite",'
                f'IF(AND(E{current_row}>=4,G{current_row}>=0.4),"Strong",'
                f'IF(N{current_row}="Yes","Playable","Thin")))'
            ))
            ws.cell(row=current_row, column=16, value=f'=E{current_row}&"-for-"&F{current_row}')
            current_row += 1

        max_row = max(ws.max_row, 1)
        if max_row > 1:
            table_ref = f"A1:P{max_row}"
            table_name = "QualifiedTable" if qualified_only else "AllMatchupsTable"
            table = Table(displayName=table_name, ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 24
        widths = {
            "A": 12,
            "B": 14,
            "C": 24,
            "D": 24,
            "E": 8,
            "F": 8,
            "G": 10,
            "H": 7,
            "I": 7,
            "J": 7,
            "K": 7,
            "L": 12,
            "M": 54,
            "N": 10,
            "O": 12,
            "P": 12,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=7).number_format = "0.0%"
            for col_idx in [5, 6, 8, 9, 10, 11, 12]:
                ws.cell(row=row_idx, column=col_idx).number_format = "0"
            ws.cell(row=row_idx, column=13).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row_idx, column=14).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=15).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=16).alignment = Alignment(horizontal="center")

        ws.conditional_formatting.add(
            f"N2:N{ws.max_row}",
            CellIsRule(operator="equal", formula=["TRUE"], fill=GREEN_FILL),
        )
        ws.conditional_formatting.add(
            f"N2:N{ws.max_row}",
            CellIsRule(operator="equal", formula=["FALSE"], fill=RED_FILL),
        )
        for tier, color in {"Elite": GREEN_FILL, "Strong": TEAL_FILL, "Playable": ORANGE_FILL, "Thin": RED_FILL}.items():
            ws.conditional_formatting.add(
                f"O2:O{ws.max_row}",
                CellIsRule(operator="equal", formula=[f'"{tier}"'], fill=color),
            )


def export_matchup_workbook(
    run_date: str,
    game_results: list[tuple[str, list[MatchupSummary]]],
    output_path: str | Path,
) -> Path:
    return WorkbookExporter(run_date=run_date, game_results=game_results).export(output_path)
